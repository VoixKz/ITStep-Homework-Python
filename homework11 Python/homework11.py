from openpyxl import Workbook, load_workbook
from requests import get
import json

combined_data = []
for i in range(1, 4):
    wb = load_workbook(f'homework11/file{i}.xlsx')
    ws = wb.active
    data = [cell.value for row in ws.iter_rows() for cell in row]
    combined_data.append(data)
sorted_data = sorted(combined_data, key = lambda x: max(x), reverse=True)
wb = Workbook()
ws = wb.active
for row_index, row in enumerate(sorted_data):
    for col_index, value in enumerate(row):
        ws.cell(row=row_index+1, column=col_index+1).value = value
wb.save('homework11/sorted_data.xlsx')



response = get('https://jsonplaceholder.typicode.com/todos/')
response_dict = response.json()
for i in range(0, len(response_dict), 40): #думаю что создавать 200 файлов не имеет смысла
    with open(f'homework11/file{i+1}-{i+40}.json', 'w') as file:
        for j in range(i, i+40):
            file.write(json.dumps(response_dict[j]))
            file.write('\n')